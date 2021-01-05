from openpyxl import *
import os as os
from PIL import Image
from openpyxl.drawing.image import Image
from ImageData import ImageData
import sqlite3

class Spreadsheet:
    def __init__(self, path, filename, image_path_list,db_name):
        self.path = path
        self.filename = filename
        self.workbook = Workbook()
        self.page = self.workbook.active
        self.create_cells()
        self.col_iter = self.page.iter_cols(min_col = 0, max_col=400, max_row=20)
        self.out_hunk(image_path_list)
        self.largest_areas = []
        this.db = db_ref
        self.save_workbook()

    def create_cells(self):
        for x in range(1, 500):
            for y in range(1, 500):
                self.page.cell(row=x, column=y)

    def out_hunk(self,image_path_list):
        # row =0
        for im_path in image_path_list:
            self.col_iter = self.page.iter_cols(min_col=1,max_col=400,min_row=row,max_row=row+20)
            self.output_set(im_path)
            row = row + 20

    def output_set(self,image_path):
        th = 73
        while(th<144):
            image_data= ImageData("frame",image_path,th,(1,5000,1),self.path)
            self.output_image(image_data)
            th= th + 5



    def output_image(self,imageData):
        seg = self.col_iter.__next__()
        headers = ["Image Name", "Porosity", "threshold","Areas"]
        values = [imageData.image_path, imageData.porosity, imageData.threshold]
        i=0
        #write headers in column
        while (i<len(headers)):
            self.page.cell(column= seg[i].column, row=seg[i].row, value = headers[i])
            i=i+1
        j=0;
        #print areas
        while(i < len(headers) + 3):
            self.page.cell(column=seg[i].column, row=seg[i].row, value = str(imageData.largest_areas[j][1]))
            j=j+1;
            i=i+1;

        seg = self.col_iter.__next__()
        i=0
        while(i<len(values)):
            self.page.cell(column= seg[i].column, row=seg[i].row, value = values[i])
            i=i+1

        self.page.cell(column= seg[i+1].column, row=seg[i+1].row, value = "(x,y)")
        #seg = self.col_iter.__next__()
        j=0
        i=i+1
        while(j<3):
            s = "(" + str(imageData.largest_areas[j][1]) + " , " +str(imageData.largest_areas[j][2]) +")"
            self.page.cell(column= seg[i].column, row=seg[i].row, value = s)
            self.largest_areas.append(
            j=j+1
            i = i+1

        seg = self.col_iter.__next__()
        seg = self.col_iter.__next__()
        img = Image(imageData.image_out_path + '/' + imageData.name + "thresh" + str(imageData.threshold) + "_out.png")
        img.width = 400
        img.height = 300
        img.anchor = str(seg[1].column_letter) +str(seg[1].row)
        self.page.add_image(img)
        i=0
        while (i<6):
            seg = self.col_iter.__next__()
            i=i+1

    def save_workbook(self):
        print("sp save")
        self.workbook.save(self.path + "/" + self.filename + ".xlsx")

    def output_frame(image_path_list, col_itter, frame_num, threshhold,color):
        i=0;
        while(i<10):
            seg = self.col_itter.__next__()
            i=i+1;

        frame_report['A' + str(frame_num)] = "frame"
        frame_report['A' + str(frame_num + 1)] = frame_name
        frame_report['B' + str(frame_num)] = "avg porosity"
        frame_report["D" + str(frame_num)] = "image"
        frame_report["E" + str(frame_num)] = "porosity"
        frame_report["F" + str(frame_num)] = "max pore"
        frame_report["G" + str(frame_num)] = "X"
        frame_report["H" + str(frame_num)] = "Y"
        for imageData in imageData_ls:
            outp


    def adapt_array(arr):
        """
http: // stackoverflow.com / a / 31312102 / 190597(SoulNibbler)
"""
out = io.BytesIO()
np.save(out, arr)
out.seek(0)
return sqlite3.Binary(out.read())






    def save_data(self):
        conn = sqlite3.connect(":memory:",detect_types=sqlite.PARSER_DECLTYPES self.db_name)
        sqlite3.register_adapter(np.ndarray, adapt_array)
        # Converts TEXT to np.array when selecting
        sqlite3.register_converter("array", convert_array)
        # Converts np.array to TEXT when inserting
        self.db.execute('insert into images (image_name,image_path,opperation_type,porosity,area_sizes)
        (?)'', (self.image_path,self.image_path,self.image_path,self.porosity,[0]))


def convert_array(text):
    out = io.BytesIO(text)
    out.seek(0)
    return np.load(out)
